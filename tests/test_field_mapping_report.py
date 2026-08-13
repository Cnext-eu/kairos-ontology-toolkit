# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Cnext.eu
"""Tests for the field-mapping report (ontology fields x one source system)."""

import textwrap

from click.testing import CliRunner
from openpyxl import load_workbook

from kairos_ontology.cli.main import cli
from kairos_ontology.core.field_mapping_report import (
    PatternConcept,
    _build_core_concepts,
    _collect_pattern_examples,
    _discover_patterns_root,
    _domain_properties,
    _extract_pattern_summary,
    _fact_source_column_uris,
    run_field_mapping_report,
    write_field_mapping_workbook,
)
from kairos_ontology.core.projections.dbt.mapping_specs import (
    AuthoredCaseBranchFact,
    AuthoredExpressionFact,
    ColumnMappingFact,
)


PARTY_TTL = textwrap.dedent("""
    @prefix party: <https://example.test/party#> .
    @prefix owl: <http://www.w3.org/2002/07/owl#> .
    @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
    @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
    <https://example.test/party> a owl:Ontology ; owl:versionInfo "1.0.0" .
    party:Customer a owl:Class ; rdfs:label "Customer" .
    party:customerId a owl:DatatypeProperty ;
      rdfs:domain party:Customer ; rdfs:range xsd:string ;
      rdfs:comment "Stable source key of the customer record." .
    party:customerName a owl:DatatypeProperty ;
      rdfs:domain party:Customer ; rdfs:range xsd:string ;
      rdfs:label "customer name" .
    party:loyaltyTier a owl:DatatypeProperty ;
      rdfs:domain party:Customer ; rdfs:range xsd:string .
    party:recordSource a owl:DatatypeProperty ;
      rdfs:domain party:Customer ; rdfs:range xsd:string .
    """).strip()


def _write_hub(hub_root):
    """A minimal v5 hub with two source systems (crm, erp) bound into one domain."""
    ontology_dir = hub_root / "model" / "ontologies"
    crm_dir = hub_root / "integration" / "sources" / "crm"
    erp_dir = hub_root / "integration" / "sources" / "erp"
    binding_dir = hub_root / "integration" / "bindings"
    ontology_dir.mkdir(parents=True)
    crm_dir.mkdir(parents=True)
    erp_dir.mkdir(parents=True)
    binding_dir.mkdir(parents=True)
    (hub_root / "kairos.yaml").write_text("adapter: fabric\n", encoding="utf-8")
    (ontology_dir / "party.ttl").write_text(PARTY_TTL, encoding="utf-8")

    (crm_dir / "crm.vocabulary.ttl").write_text(
        textwrap.dedent("""
            @prefix src: <https://example.test/source/crm#> .
            @prefix kb: <https://kairos.cnext.eu/bronze#> .
            @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
            @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
            src:crm a kb:SourceSystem ; rdfs:label "crm" .
            src:customers a kb:SourceTable ; kb:sourceSystem src:crm ;
              kb:tableName "customers" ; kb:primaryKeyColumns "customer_id" .
            src:id a kb:SourceColumn ; kb:sourceTable src:customers ;
              kb:columnName "customer_id" ; kb:dataType "varchar(50)" ;
              kb:nullable "false"^^xsd:boolean ; kb:sampleValues "C-1 | C-2" .
            src:name a kb:SourceColumn ; kb:sourceTable src:customers ;
              kb:columnName "customer_name" ; kb:dataType "varchar(200)" ;
              kb:nullable "true"^^xsd:boolean ; kb:sampleValues "Acme | Globex" .
            """).strip(),
        encoding="utf-8",
    )
    (erp_dir / "erp.vocabulary.ttl").write_text(
        textwrap.dedent("""
            @prefix src: <https://example.test/source/erp#> .
            @prefix kb: <https://kairos.cnext.eu/bronze#> .
            @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
            @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
            src:erp a kb:SourceSystem ; rdfs:label "erp" .
            src:accounts a kb:SourceTable ; kb:sourceSystem src:erp ;
              kb:tableName "accounts" ; kb:primaryKeyColumns "acct_id" .
            src:id a kb:SourceColumn ; kb:sourceTable src:accounts ;
              kb:columnName "acct_id" ; kb:dataType "varchar(50)" ;
              kb:nullable "false"^^xsd:boolean ; kb:sampleValues "A-1 | A-2" .
            src:name a kb:SourceColumn ; kb:sourceTable src:accounts ;
              kb:columnName "acct_name" ; kb:dataType "varchar(200)" ;
              kb:nullable "true"^^xsd:boolean ; kb:sampleValues "Acme Corp" .
            """).strip(),
        encoding="utf-8",
    )
    (binding_dir / "crm-customer.binding.yaml").write_text(
        textwrap.dedent("""
            apiVersion: kairos.eu/v5
            kind: EntityBinding
            metadata:
              name: crm-customer
              domain: party
            source:
              relation: crm.customers
            target:
              class: party:Customer
            grain:
              columns: [customer_id]
            identity:
              strategy: source-natural
              sourceKey: [customer_id]
            load:
              mode: full-refresh
            fields:
              - property: party:customerId
                expression: customer_id
              - property: party:customerName
                expression: customer_name
              - property: party:recordSource
                expression: { literal: "crm", datatype: "string" }
            """).strip(),
        encoding="utf-8",
    )
    (binding_dir / "erp-account.binding.yaml").write_text(
        textwrap.dedent("""
            apiVersion: kairos.eu/v5
            kind: EntityBinding
            metadata:
              name: erp-account
              domain: party
            source:
              relation: erp.accounts
            target:
              class: party:Customer
            grain:
              columns: [acct_id]
            identity:
              strategy: source-natural
              sourceKey: [acct_id]
            load:
              mode: full-refresh
            fields:
              - property: party:customerName
                expression: acct_name
            technicalFields:
              - name: acct_id
                expression: acct_id
                type: string
                nullable: false
                purpose: identity
            """).strip(),
        encoding="utf-8",
    )
    return ontology_dir, binding_dir, hub_root / "integration" / "sources"


def test_domain_properties_prefers_comment_over_label_and_reads_domain_class(tmp_path):
    ontology_path = tmp_path / "party.ttl"
    ontology_path.write_text(PARTY_TTL, encoding="utf-8")

    rows, notes = _domain_properties(ontology_path)
    by_local = {meta.property_local: meta for meta in rows}

    assert notes == []
    assert by_local["customerId"].class_local == "Customer"
    assert by_local["customerId"].description == "Stable source key of the customer record."
    assert by_local["customerId"].range_display == "xsd:string"
    assert by_local["customerId"].origin == "direct"
    assert by_local["customerName"].description == "customer name"
    assert by_local["loyaltyTier"].description == ""


SUBCLASS_TTL = textwrap.dedent("""
    @prefix party: <https://example.test/party2#> .
    @prefix owl: <http://www.w3.org/2002/07/owl#> .
    @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
    @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
    <https://example.test/party2> a owl:Ontology ; owl:versionInfo "1.0.0" .
    party:TradeParty a owl:Class .
    party:Customer a owl:Class ; rdfs:subClassOf party:TradeParty .
    party:partyName a owl:DatatypeProperty ;
      rdfs:domain party:TradeParty ; rdfs:range xsd:string ;
      rdfs:comment "Name of the trade party." .
    party:customerId a owl:DatatypeProperty ;
      rdfs:domain party:Customer ; rdfs:range xsd:string .
    """).strip()


def test_domain_properties_includes_same_file_inherited_properties(tmp_path):
    ontology_path = tmp_path / "party2.ttl"
    ontology_path.write_text(SUBCLASS_TTL, encoding="utf-8")

    rows, notes = _domain_properties(ontology_path)
    by_class_field = {(meta.class_local, meta.property_local): meta for meta in rows}

    assert notes == []
    assert by_class_field[("Customer", "customerId")].origin == "direct"
    # partyName is declared on the ancestor TradeParty, not Customer itself -- must still
    # show up under Customer (tagged "inherited"), which is the whole point of this fix.
    assert by_class_field[("Customer", "partyName")].origin == "inherited"
    assert by_class_field[("Customer", "partyName")].description == "Name of the trade party."
    # TradeParty's own tab-row for partyName is unaffected.
    assert by_class_field[("TradeParty", "partyName")].origin == "direct"


def test_domain_properties_includes_properties_inherited_across_owl_imports(tmp_path):
    """The actual crux of the owl:imports question: a class subclassing an ancestor
    declared in a DIFFERENT, imported file must still surface the ancestor's properties
    under the concrete class -- not just properties asserted in the domain's own file."""
    (tmp_path / "foundation.ttl").write_text(
        textwrap.dedent("""
            @prefix found: <https://example.test/foundation#> .
            @prefix owl: <http://www.w3.org/2002/07/owl#> .
            @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
            @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
            <https://example.test/foundation> a owl:Ontology ; owl:versionInfo "1.0.0" .
            found:TradeParty a owl:Class .
            found:partyName a owl:DatatypeProperty ;
              rdfs:domain found:TradeParty ; rdfs:range xsd:string ;
              rdfs:comment "Name of the trade party." .
            """).strip(),
        encoding="utf-8",
    )
    (tmp_path / "party3.ttl").write_text(
        textwrap.dedent("""
            @prefix party: <https://example.test/party3#> .
            @prefix found: <https://example.test/foundation#> .
            @prefix owl: <http://www.w3.org/2002/07/owl#> .
            @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
            @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
            <https://example.test/party3> a owl:Ontology ; owl:versionInfo "1.0.0" ;
              owl:imports <https://example.test/foundation> .
            party:Customer a owl:Class ; rdfs:subClassOf found:TradeParty .
            party:customerId a owl:DatatypeProperty ;
              rdfs:domain party:Customer ; rdfs:range xsd:string .
            """).strip(),
        encoding="utf-8",
    )
    (tmp_path / "catalog-v001.xml").write_text(
        "<?xml version='1.0'?>\n"
        '<catalog xmlns="urn:oasis:names:tc:entity:xmlns:xml:catalog">\n'
        '  <uri name="https://example.test/foundation" uri="foundation.ttl"/>\n'
        "</catalog>\n",
        encoding="utf-8",
    )

    rows, notes = _domain_properties(tmp_path / "party3.ttl")
    by_class_field = {(meta.class_local, meta.property_local): meta for meta in rows}

    assert notes == []
    assert by_class_field[("Customer", "customerId")].origin == "direct"
    assert by_class_field[("Customer", "partyName")].origin == "inherited"
    assert by_class_field[("Customer", "partyName")].description == "Name of the trade party."
    # TradeParty itself isn't asserted in party3.ttl (only imported) -- it must not get
    # its own tab-row here; the report's own domain file is the only class-enumeration root.
    assert ("TradeParty", "partyName") not in by_class_field


def test_fact_source_column_uris_returns_direct_reference():
    fact = ColumnMappingFact(
        resource_uri="urn:m1",
        source_column_uri="table/col",
        target_property_uri="urn:p1",
        match_type="exactMatch",
    )
    assert _fact_source_column_uris(fact) == ("table/col",)


def test_fact_source_column_uris_walks_compound_case_expression():
    left = AuthoredExpressionFact(
        resource_uri="urn:e1",
        kind="source-column",
        output_type="string",
        nullable="true",
        null_policy="propagate",
        determinism="deterministic",
        capabilities=("source-column",),
        source_column_uri="table/a",
    )
    right = AuthoredExpressionFact(
        resource_uri="urn:e2",
        kind="source-column",
        output_type="string",
        nullable="true",
        null_policy="propagate",
        determinism="deterministic",
        capabilities=("source-column",),
        source_column_uri="table/b",
    )
    case_expr = AuthoredExpressionFact(
        resource_uri="urn:e3",
        kind="case",
        output_type="string",
        nullable="true",
        null_policy="propagate",
        determinism="deterministic",
        capabilities=("case-expression",),
        branches=(AuthoredCaseBranchFact(resource_uri="urn:b1", condition=left, result=right),),
    )
    fact = ColumnMappingFact(
        resource_uri="urn:m2",
        source_column_uri="",
        target_property_uri="urn:p2",
        match_type="exactMatch",
        expression=case_expr,
    )
    assert set(_fact_source_column_uris(fact)) == {"table/a", "table/b"}


def test_run_field_mapping_report_maps_bindings_with_sample_values(tmp_path):
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)

    report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
    )

    rows = {row.property_local: row for row in report.rows_by_domain["party"]}
    assert rows["customerId"].source_columns == ("customers.customer_id",)
    assert rows["customerId"].sample_value == "C-1"
    assert rows["customerName"].source_columns == ("customers.customer_name",)
    assert rows["customerName"].sample_value == "Acme"


def test_run_field_mapping_report_filters_by_source_system(tmp_path):
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)

    crm_report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
    )
    erp_report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="erp",
    )

    crm_rows = {row.property_local: row for row in crm_report.rows_by_domain["party"]}
    erp_rows = {row.property_local: row for row in erp_report.rows_by_domain["party"]}

    # customerName is bound from both systems -- crm's report must show only its own
    # source column, not erp's, and vice versa.
    assert crm_rows["customerName"].source_columns == ("customers.customer_name",)
    assert erp_rows["customerName"].source_columns == ("accounts.acct_name",)
    # customerId is only ever bound from crm.
    assert erp_rows["customerId"].source_columns == ()
    assert erp_rows["customerId"].sample_value == ""


def test_run_field_mapping_report_shows_unmapped_fields_not_hidden(tmp_path):
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)

    report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
    )

    rows = {row.property_local: row for row in report.rows_by_domain["party"]}
    assert "loyaltyTier" in rows
    assert rows["loyaltyTier"].source_columns == ()
    assert rows["loyaltyTier"].sample_value == ""


def test_run_field_mapping_report_literal_expression_is_not_falsely_attributed(tmp_path):
    """A constant-valued ``fields:`` mapping must show as unmapped, never as sourced from
    the arbitrary anchor column ``adapter.py``'s ``_build_field_mappings`` assigns a
    leafless expression's ``ColumnMappingFact.source_column_uri`` for its own bookkeeping."""
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)

    report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
    )

    rows = {row.property_local: row for row in report.rows_by_domain["party"]}
    assert rows["recordSource"].source_columns == ()
    assert rows["recordSource"].sample_value == ""


def test_run_field_mapping_report_dedups_two_bindings_same_system(tmp_path):
    """Two DIFFERENT bindings under the SAME source system mapping the same target
    property must both show up (deduplicated, not doubled) rather than one silently
    winning -- the crm/erp fixture only exercises cross-system dedup, not this."""
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)

    (sources_dir / "crm" / "crm-leads.vocabulary.ttl").write_text(
        textwrap.dedent("""
            @prefix src: <https://example.test/source/crm2#> .
            @prefix kb: <https://kairos.cnext.eu/bronze#> .
            @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
            @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
            src:crm a kb:SourceSystem ; rdfs:label "crm" .
            src:leads a kb:SourceTable ; kb:sourceSystem src:crm ;
              kb:tableName "leads" ; kb:primaryKeyColumns "lead_id" .
            src:id a kb:SourceColumn ; kb:sourceTable src:leads ;
              kb:columnName "lead_id" ; kb:dataType "varchar(50)" ;
              kb:nullable "false"^^xsd:boolean ; kb:sampleValues "L-1" .
            src:name a kb:SourceColumn ; kb:sourceTable src:leads ;
              kb:columnName "lead_name" ; kb:dataType "varchar(200)" ;
              kb:nullable "true"^^xsd:boolean ; kb:sampleValues "Prospect Co" .
            """).strip(),
        encoding="utf-8",
    )
    (binding_dir / "crm-lead.binding.yaml").write_text(
        textwrap.dedent("""
            apiVersion: kairos.eu/v5
            kind: EntityBinding
            metadata:
              name: crm-lead
              domain: party
            source:
              relation: crm.leads
            target:
              class: party:Customer
            grain:
              columns: [lead_id]
            identity:
              strategy: source-natural
              sourceKey: [lead_id]
            load:
              mode: full-refresh
            fields:
              - property: party:customerId
                expression: lead_id
              - property: party:customerName
                expression: lead_name
            """).strip(),
        encoding="utf-8",
    )

    report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
    )

    rows = {row.property_local: row for row in report.rows_by_domain["party"]}
    # "crm-customer.binding.yaml" sorts before "crm-lead.binding.yaml", so
    # customers.customer_name resolves first and its sample wins deterministically.
    assert rows["customerName"].source_columns == (
        "customers.customer_name",
        "leads.lead_name",
    )
    assert rows["customerName"].sample_value == "Acme"


def test_run_field_mapping_report_domain_filter(tmp_path):
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)

    report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
        domains=("party",),
    )

    assert set(report.rows_by_domain) == {"party"}


def test_write_field_mapping_workbook_creates_cover_and_domain_sheets(tmp_path):
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)
    report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
    )

    output_path = tmp_path / "out" / "field-mapping-crm.xlsx"
    write_field_mapping_workbook(report, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Overview", "party"]
    sheet = workbook["party"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "Ontology Class (Label)",
        "Ontology Field (Label)",
        "Origin",
        "Ontology Description",
        "Range",
        "Source Field(s)",
        "Source Field Example",
        "Ontology Reference IRI",
    ]
    # Header row is styled: bold white text on a colored fill, frozen, auto-filtered.
    header_cell = next(sheet.iter_rows(min_row=1, max_row=1))[0]
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb == "001F4E78"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None
    # Columns are auto-sized (not the openpyxl default of None/8.43).
    assert sheet.column_dimensions["A"].width is not None

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    by_field = {row[1]: row for row in rows}
    assert by_field["customerId"][2] == "direct"
    assert by_field["customerId"][4] == "xsd:string"
    assert by_field["customerId"][5] == "customers.customer_id"
    assert by_field["customerId"][6] == "C-1"
    assert by_field["loyaltyTier"][5] == "NO-MAPPING-FOUND"
    # openpyxl round-trips an empty-string cell as None, not "".
    assert by_field["loyaltyTier"][6] is None


def test_cli_field_mapping_report_writes_workbook(tmp_path):
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)
    output_path = tmp_path / "out" / "field-mapping-crm.xlsx"
    runner = CliRunner()

    result = runner.invoke(
        cli,
        [
            "field-mapping-report",
            "--ontologies",
            str(ontology_dir),
            "--bindings",
            str(binding_dir),
            "--sources",
            str(sources_dir),
            "--source-system",
            "crm",
            "--output",
            str(output_path),
        ],
    )

    assert result.exit_code == 0, result.output
    assert output_path.is_file()
    assert "1 domain(s)" in result.output


_PATTERN_TTL = textwrap.dedent("""
    @prefix cons: <https://example.test/cons#> .
    @prefix owl: <http://www.w3.org/2002/07/owl#> .
    @prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
    @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
    <https://example.test/cons> a owl:Ontology ; owl:versionInfo "1.0.0" .
    cons:Shipment a owl:Class ; rdfs:label "Shipment" .
    cons:hasMasterWaybill a owl:ObjectProperty ;
      rdfs:domain cons:Shipment ;
      rdfs:comment "The master waybill covering this shipment. Eventual object property under patterns/deferred-relationship: the documents domain is not yet onboarded." .
    """).strip()

_DEFERRED_RELATIONSHIP_PATTERN_MD = textwrap.dedent("""
    # Deferred Relationship

    **Normativity:** naming — normative.

    ## Problem

    A domain slice needs to land before the cross-domain key it should link to has
    conformed. This is the first paragraph of the problem statement.

    Second paragraph should not be included.

    ## Applicability

    Not part of the explanation.
    """).strip()


def _write_pattern_library(root, pattern_id="deferred-relationship", markdown=_DEFERRED_RELATIONSHIP_PATTERN_MD):
    pattern_dir = root / "ontology-reference-models" / "blueprints" / "patterns" / pattern_id
    pattern_dir.mkdir(parents=True)
    (pattern_dir / "pattern.md").write_text(markdown, encoding="utf-8")


def test_collect_pattern_examples_finds_object_property_reference(tmp_path):
    ontology_path = tmp_path / "consignment.ttl"
    ontology_path.write_text(_PATTERN_TTL, encoding="utf-8")

    examples = _collect_pattern_examples([ontology_path])

    assert set(examples) == {"deferred-relationship"}
    example = examples["deferred-relationship"]
    assert example.domain == "consignment"
    assert example.class_local == "Shipment"
    assert example.property_local == "hasMasterWaybill"
    assert example.excerpt.startswith("The master waybill covering this shipment.")


def test_discover_patterns_root_checks_hub_root_then_sibling(tmp_path):
    hub_root = tmp_path / "some-hub" / "ontology-hub"
    hub_root.mkdir(parents=True)

    assert _discover_patterns_root(hub_root) is None

    _write_pattern_library(hub_root.parent)

    found = _discover_patterns_root(hub_root)
    assert found == hub_root.parent / "ontology-reference-models" / "blueprints" / "patterns"


def test_extract_pattern_summary_reads_title_and_problem_paragraph_only():
    title, explanation = _extract_pattern_summary(_DEFERRED_RELATIONSHIP_PATTERN_MD)

    assert title == "Deferred Relationship"
    assert explanation.startswith("A domain slice needs to land")
    assert "Second paragraph" not in explanation
    assert "Not part of the explanation" not in explanation


def test_build_core_concepts_pairs_pattern_summary_with_real_example(tmp_path):
    ontology_path = tmp_path / "consignment.ttl"
    ontology_path.write_text(_PATTERN_TTL, encoding="utf-8")
    _write_pattern_library(tmp_path)

    notes: list[str] = []
    concepts = _build_core_concepts([ontology_path], tmp_path, notes)

    assert notes == []
    assert concepts == [
        PatternConcept(
            pattern_id="deferred-relationship",
            title="Deferred Relationship",
            explanation=(
                "A domain slice needs to land before the cross-domain key it should link to "
                "has conformed. This is the first paragraph of the problem statement."
            ),
            example_domain="consignment",
            example_class="Shipment",
            example_property="hasMasterWaybill",
            example_excerpt=concepts[0].example_excerpt,
        )
    ]
    assert concepts[0].example_excerpt.startswith("The master waybill covering this shipment.")


def test_build_core_concepts_degrades_gracefully_when_pattern_library_missing(tmp_path):
    ontology_path = tmp_path / "consignment.ttl"
    ontology_path.write_text(_PATTERN_TTL, encoding="utf-8")

    notes: list[str] = []
    concepts = _build_core_concepts([ontology_path], tmp_path, notes)

    assert concepts == []
    assert any("Blueprint pattern library not found" in note for note in notes)


def test_write_field_mapping_workbook_inserts_core_concepts_sheet_second(tmp_path):
    ontology_dir, binding_dir, sources_dir = _write_hub(tmp_path)
    (ontology_dir / "consignment.ttl").write_text(_PATTERN_TTL, encoding="utf-8")
    _write_pattern_library(tmp_path)

    report = run_field_mapping_report(
        ontologies_path=ontology_dir,
        bindings_dir=binding_dir,
        sources_dir=sources_dir,
        hub_root=tmp_path,
        source_system="crm",
    )
    assert len(report.core_concepts) == 1

    output_path = tmp_path / "out" / "field-mapping-crm.xlsx"
    write_field_mapping_workbook(report, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Overview", "Core Concepts", "consignment", "party"]
    concepts_sheet = workbook["Core Concepts"]
    values = [cell.value for row in concepts_sheet.iter_rows() for cell in row if cell.value]
    assert "Deferred Relationship" in values
    assert any("A domain slice needs to land" in v for v in values)
    assert any("consignment: :Shipment -> :hasMasterWaybill" in v for v in values)
