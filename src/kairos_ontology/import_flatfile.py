# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Cnext.eu
"""Import-Flatfile — create source schema YAML from CSV/Excel flat files.

Reads CSV or Excel files and produces the same intermediate YAML + samples format
that extract-schema generates from live databases. The output directory can then be
passed to ``import-source`` to generate bronze vocabulary TTL.

Pipeline: CSV/XLSX → _manifest.yaml + {table}.yaml + {table}.samples.yaml
          → import-source → .vocabulary.ttl → analyse-sources
"""

from __future__ import annotations

import csv
import logging
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml

# Increase CSV field size limit to handle large fields (e.g., Oracle exports)
csv.field_size_limit(sys.maxsize)

logger = logging.getLogger(__name__)

# Maximum rows to read for type inference (to avoid loading huge files into memory)
DEFAULT_MAX_ROWS = 1000
DEFAULT_SAMPLE_SIZE = 5


# --------------------------------------------------------------------------- #
# Type Inference
# --------------------------------------------------------------------------- #

# Date/datetime patterns
_DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),  # 2024-01-15
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),  # 15/01/2024
    re.compile(r"^\d{2}-\d{2}-\d{4}$"),  # 15-01-2024
]
_DATETIME_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}"),  # ISO datetime
]
_BOOL_VALUES = {"true", "false", "1", "0", "yes", "no", "y", "n"}


def infer_column_type(values: list[str]) -> str:
    """Infer SQL-like data type from a list of string values.

    Args:
        values: Non-empty string values from the column (nulls/blanks excluded).

    Returns:
        One of: 'bigint', 'int', 'decimal', 'date', 'datetime', 'bit', 'varchar(max)'
    """
    if not values:
        return "varchar(max)"

    # Check for boolean
    if all(v.lower().strip() in _BOOL_VALUES for v in values):
        return "bit"

    # Check for datetime (before date — datetime is a superset)
    if all(any(p.match(v.strip()) for p in _DATETIME_PATTERNS) for v in values):
        return "datetime"

    # Check for date
    if all(any(p.match(v.strip()) for p in _DATE_PATTERNS) for v in values):
        return "date"

    # Check for integer
    int_count = 0
    for v in values:
        stripped = v.strip()
        try:
            int(stripped)
            int_count += 1
        except ValueError:
            break
    if int_count == len(values):
        max_val = max(abs(int(v.strip())) for v in values)
        return "bigint" if max_val > 2_147_483_647 else "int"

    # Check for decimal/float
    float_count = 0
    for v in values:
        stripped = v.strip()
        try:
            float(stripped)
            float_count += 1
        except ValueError:
            break
    if float_count == len(values):
        return "decimal"

    return "varchar(max)"


# --------------------------------------------------------------------------- #
# CSV Reading
# --------------------------------------------------------------------------- #


def read_csv_table(
    path: Path, max_rows: int = DEFAULT_MAX_ROWS, sample_size: int = DEFAULT_SAMPLE_SIZE
) -> dict[str, Any]:
    """Read a CSV file and produce a table data dict.

    Args:
        path: Path to the .csv file.
        max_rows: Maximum rows to read for type inference.
        sample_size: Number of sample rows to store.

    Returns:
        Dict with keys: name, row_count, columns, sample_rows.
    """
    table_name = path.stem

    with open(path, encoding="utf-8-sig", newline="") as f:
        # Sniff the dialect
        sample_text = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample_text, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(f, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError(f"No headers found in CSV file: {path}")

        headers = list(reader.fieldnames)
        all_rows: list[dict[str, str]] = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            all_rows.append(row)

    row_count = len(all_rows)

    # Build column metadata
    columns = []
    for pos, col_name in enumerate(headers, start=1):
        non_empty_values = [
            (row.get(col_name) or "").strip()
            for row in all_rows
            if (row.get(col_name) or "").strip()
        ]
        distinct_values = list(dict.fromkeys(non_empty_values))
        nullable = len(non_empty_values) < row_count

        col_dict: dict[str, Any] = {
            "name": col_name,
            "data_type": infer_column_type(distinct_values[:100]),
            "ordinal_position": pos,
            "nullable": nullable,
        }
        if distinct_values:
            col_dict["distinct_count"] = len(distinct_values)
            col_dict["samples"] = distinct_values[:sample_size]

        columns.append(col_dict)

    # Sample rows (raw dicts for .samples.yaml)
    sample_rows = [
        {k: v for k, v in row.items() if v}
        for row in all_rows[:sample_size]
    ]

    return {
        "name": table_name,
        "row_count": row_count,
        "columns": columns,
        "sample_rows": sample_rows,
    }


# --------------------------------------------------------------------------- #
# Excel Reading
# --------------------------------------------------------------------------- #


def read_xlsx_tables(
    path: Path, max_rows: int = DEFAULT_MAX_ROWS, sample_size: int = DEFAULT_SAMPLE_SIZE
) -> list[dict[str, Any]]:
    """Read an Excel workbook and produce one table dict per worksheet.

    Requires openpyxl (install via: pip install kairos-ontology-toolkit[flatfile]).

    Args:
        path: Path to the .xlsx file.
        max_rows: Maximum rows to read per sheet for type inference.
        sample_size: Number of sample rows to store.

    Returns:
        List of table data dicts (same format as read_csv_table).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel support. "
            "Install with: pip install kairos-ontology-toolkit[flatfile]"
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # First row = headers
        try:
            headers_row = next(rows_iter)
        except StopIteration:
            continue

        headers = [str(h).strip() if h else f"Column{i}" for i, h in enumerate(headers_row, 1)]
        if not any(h for h in headers):
            continue

        # Read data rows
        all_rows: list[dict[str, str]] = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                break
            row_dict = {}
            for col_name, val in zip(headers, row):
                row_dict[col_name] = str(val) if val is not None else ""
            all_rows.append(row_dict)

        row_count = len(all_rows)
        if row_count == 0:
            continue

        # Build column metadata
        columns = []
        for pos, col_name in enumerate(headers, start=1):
            non_empty_values = [
                (row.get(col_name) or "").strip()
                for row in all_rows
                if (row.get(col_name) or "").strip()
            ]
            distinct_values = list(dict.fromkeys(non_empty_values))
            nullable = len(non_empty_values) < row_count

            col_dict: dict[str, Any] = {
                "name": col_name,
                "data_type": infer_column_type(distinct_values[:100]),
                "ordinal_position": pos,
                "nullable": nullable,
            }
            if distinct_values:
                col_dict["distinct_count"] = len(distinct_values)
                col_dict["samples"] = distinct_values[:sample_size]

            columns.append(col_dict)

        sample_rows = [
            {k: v for k, v in row.items() if v}
            for row in all_rows[:sample_size]
        ]

        tables.append({
            "name": sheet_name,
            "row_count": row_count,
            "columns": columns,
            "sample_rows": sample_rows,
        })

    wb.close()
    return tables


# --------------------------------------------------------------------------- #
# Output Writing
# --------------------------------------------------------------------------- #


def write_source_dir(
    tables: list[dict[str, Any]],
    system_name: str,
    output_dir: Path,
    platform: str = "flatfile",
) -> Path:
    """Write table data to the standard source directory format.

    Creates:
      - _manifest.yaml
      - {table}.yaml per table (schema metadata)
      - {table}.samples.yaml per table (sample rows)

    Args:
        tables: List of table data dicts (from read_csv_table or read_xlsx_tables).
        system_name: System name for the source.
        output_dir: Target directory for output files.
        platform: Platform identifier (default: "flatfile").

    Returns:
        Path to the output directory.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Write manifest
    manifest = {
        "version": "1.1",
        "system": system_name,
        "platform": platform,
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "connection": {},
        "tables": [t["name"] for t in tables],
    }
    with open(output_dir / "_manifest.yaml", "w", encoding="utf-8") as f:
        yaml.dump(manifest, f, default_flow_style=False, sort_keys=False)

    # Write per-table YAML + samples
    for table in tables:
        tbl_name = table["name"]

        # Schema metadata (without sample_rows — those go in .samples.yaml)
        table_yaml = {
            "name": tbl_name,
            "row_count": table.get("row_count", 0),
            "columns": [
                {k: v for k, v in col.items() if k != "samples"}
                for col in table["columns"]
            ],
        }
        with open(output_dir / f"{tbl_name}.yaml", "w", encoding="utf-8") as f:
            yaml.dump(table_yaml, f, default_flow_style=False, sort_keys=False)

        # Sample rows
        sample_rows = table.get("sample_rows", [])
        if sample_rows:
            samples_data = {
                "extracted_at": manifest["extracted_at"],
                "table": tbl_name,
                "schema": "",
                "rows": sample_rows,
            }
            with open(output_dir / f"{tbl_name}.samples.yaml", "w", encoding="utf-8") as f:
                yaml.dump(samples_data, f, default_flow_style=False, sort_keys=False)

    logger.info("Written %d table(s) to %s", len(tables), output_dir)
    return output_dir


# --------------------------------------------------------------------------- #
# Main Orchestration
# --------------------------------------------------------------------------- #


def run_import_flatfile(
    source_path: Path,
    system_name: str | None = None,
    output_dir: Path | None = None,
    max_rows: int = DEFAULT_MAX_ROWS,
    sample_size: int = DEFAULT_SAMPLE_SIZE,
) -> Path:
    """Orchestrate the flatfile import workflow.

    Accepts a single CSV, single XLSX, or a directory containing CSV/XLSX files.

    Args:
        source_path: Path to CSV, XLSX file, or directory.
        system_name: Override system name (default: derived from path).
        output_dir: Output directory (default: integration/sources/{system}/).
        max_rows: Maximum rows to read for type inference.
        sample_size: Number of sample rows to store.

    Returns:
        Path to the output directory.
    """
    tables: list[dict[str, Any]] = []

    if source_path.is_file():
        suffix = source_path.suffix.lower()
        default_name = source_path.stem

        if suffix == ".csv":
            tables.append(read_csv_table(source_path, max_rows, sample_size))
        elif suffix in (".xlsx", ".xls"):
            tables.extend(read_xlsx_tables(source_path, max_rows, sample_size))
        else:
            raise ValueError(
                f"Unsupported file type: {suffix}. Use .csv or .xlsx"
            )

    elif source_path.is_dir():
        default_name = source_path.name
        files = sorted(source_path.iterdir())
        for f in files:
            if f.suffix.lower() == ".csv":
                tables.append(read_csv_table(f, max_rows, sample_size))
            elif f.suffix.lower() in (".xlsx", ".xls"):
                tables.extend(read_xlsx_tables(f, max_rows, sample_size))

        if not tables:
            raise ValueError(
                f"No CSV or Excel files found in: {source_path}"
            )
    else:
        raise ValueError(f"Path does not exist: {source_path}")

    if not system_name:
        system_name = default_name

    if output_dir is None:
        # Detect hub root
        cwd = Path.cwd()
        hub_root = None
        for candidate in [cwd / "ontology-hub", cwd]:
            if (candidate / "model" / "ontologies").is_dir():
                hub_root = candidate
                break
        if hub_root:
            output_dir = hub_root / "integration" / "sources" / system_name
        else:
            logger.warning(
                "Could not detect ontology-hub root (no model/ontologies/ found). "
                "Writing to relative path: integration/sources/%s. "
                "Use --output to specify an explicit output directory.",
                system_name,
            )
            output_dir = Path("integration/sources") / system_name

    return write_source_dir(tables, system_name, output_dir)
