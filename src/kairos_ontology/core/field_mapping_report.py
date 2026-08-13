# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Cnext.eu
"""Field-mapping report: ontology fields cross-referenced against one source system.

For each domain ontology, lists every ``owl:DatatypeProperty`` that applies to a class the
domain declares -- both properties declared directly on the class ("direct") and ones
inherited from an ancestor class, whether that ancestor lives in the same file or an
``owl:imports``-ed foundation/reference module ("inherited") -- alongside its
ontology-authored description and IRI, then joins it against the EntityBindings that map a
chosen source system (e.g. ``cargowise``) onto that property -- embedding the mapped source
column and a real sample value when source vocabulary/sample data is available.

Object properties (relationships) are out of scope for this report: EntityBinding
``relationships:`` joins are wired later in the compile pipeline (``_wire_relationships``,
``core/compiler/kernel.py``) and don't carry a single representative "value" the way a
scalar ``fields:`` mapping does. Only ``fields:``-declared mappings are shown.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from .hub_utils import is_domain_ontology_stem
from .ontology_loader import SemanticProfile, load_ontology
from .projections.dbt.mapping_bind import expression_input_uris
from .projections.dbt.mapping_specs import ColumnMappingFact
from .projections.uri_utils import extract_local_name
from .silver_sample_audit import (
    SourceColumnSample,
    _resolve_source_column,
    load_source_samples,
    resolve_v5_column_facts,
)


@dataclass
class FieldMappingRow:
    """One ontology field's mapping to the selected source system, if any."""

    domain: str
    class_local: str
    property_local: str
    property_uri: str
    description: str
    range_display: str = ""
    origin: str = "direct"
    source_columns: tuple[str, ...] = ()
    sample_value: str = ""


@dataclass
class _PropertyMeta:
    """One domain-asserted ``owl:DatatypeProperty``, ready to place in a report row."""

    class_local: str
    property_local: str
    property_uri: str
    description: str
    range_display: str
    origin: str = "direct"


_XSD_NS = "http://www.w3.org/2001/XMLSchema#"


def _compact_datatype_uri(uri: str) -> str:
    """Return a short, readable form of a datatype URI (``xsd:string``), not the full IRI."""
    if uri.startswith(_XSD_NS):
        return f"xsd:{uri[len(_XSD_NS):]}"
    return extract_local_name(uri)


@dataclass
class FieldMappingReport:
    """Structured output of the field-mapping report."""

    source_system: str
    generated_at: str
    ontologies_path: str
    bindings_dir: str
    sources_dir: str
    rows_by_domain: dict[str, list[FieldMappingRow]] = field(default_factory=dict)
    core_concepts: list["PatternConcept"] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def _property_description(record) -> str:
    # SemanticIndex defaults an undeclared rdfs:label to the term's own local name
    # (semantic_index.py), so it's never empty -- only trust it as a real description
    # when it differs from the local name; otherwise fall through to "" rather than
    # echoing the field's own name back as its "description".
    authored_label = record.label if record.label != record.name else ""
    return record.comment or authored_label


def _domain_properties(ontology_path: Path) -> tuple[list[_PropertyMeta], list[str]]:
    """Return metadata for every ``owl:DatatypeProperty`` that applies to a class asserted
    directly in *ontology_path* (not properties pulled in wholesale through unrelated
    ``owl:imports``).

    Routed through the canonical ``load_ontology``/``SemanticIndex`` loader (DD-103)
    rather than a direct ``rdflib`` parse, so prefix/import resolution matches every other
    domain-facing command (e.g. ``explain-term``), using the ``rdfs`` profile so
    ``ClassRecord.inherited_properties`` is actually computed (the default ``asserted``
    profile skips subclass-transitivity entirely).

    Two passes, both scoped to *this domain's own classes/properties*
    (``provenance.import_depth == 0``), not imported ones in general:

    1. For every class this domain declares directly, ``SemanticIndex.class_properties``
       returns its ``direct`` properties (declared with ``rdfs:domain`` on the class
       itself) *and* its ``inherited`` ones (declared on an ancestor class -- whether that
       ancestor lives in this same file or a ``owl:imports``-ed foundation/reference
       module). A class genuinely inheriting shared fields from a foundation class would
       otherwise show none of them, which is the gap this fixes.
    2. A property directly asserted in this file whose declared domain class is *not*
       one of this domain's own classes (e.g. a domain pointing at an imported class
       without redeclaring it) would be invisible to pass 1 -- kept as a fallback so pass 1
       can only add coverage, never drop something the previous single-pass walk caught.
    """
    notes: list[str] = []
    loaded = load_ontology(ontology_path, degraded=True, profile=SemanticProfile.RDFS)
    if not loaded.complete:
        notes.append(
            f"'{ontology_path.name}': import closure is incomplete; "
            "some inherited/imported context may be missing from its property list."
        )

    index = loaded.semantic_index
    rows: list[_PropertyMeta] = []
    seen: set[tuple[str, str]] = set()  # (class_uri, property_uri)

    for cls in index.classes:
        if cls.provenance.import_depth != 0:
            continue
        class_local = extract_local_name(cls.uri)
        for entry in index.class_properties(cls.uri):
            if entry["property_type"] != "datatype":
                continue
            key = (cls.uri, entry["property_uri"])
            if key in seen:
                continue
            seen.add(key)
            record = index.property_by_uri(entry["property_uri"])
            description = _property_description(record) if record else ""
            range_display = _compact_datatype_uri(entry["ranges"][0]) if entry["ranges"] else ""
            rows.append(
                _PropertyMeta(
                    class_local=class_local,
                    property_local=entry["name"],
                    property_uri=entry["property_uri"],
                    description=description,
                    range_display=range_display,
                    origin=entry["origin"],
                )
            )

    for record in index.properties:
        if record.property_type != "datatype" or record.provenance.import_depth != 0:
            continue
        domain_uris = tuple(link.uri for link in record.domains) or ("",)
        for class_uri in domain_uris:
            key = (class_uri, record.uri)
            if key in seen:
                continue
            seen.add(key)
            class_local = extract_local_name(class_uri) if class_uri else ""
            range_display = _compact_datatype_uri(record.ranges[0].uri) if record.ranges else ""
            rows.append(
                _PropertyMeta(
                    class_local=class_local,
                    property_local=record.name,
                    property_uri=record.uri,
                    description=_property_description(record),
                    range_display=range_display,
                    origin="direct",
                )
            )

    rows.sort(key=lambda r: (r.class_local, r.origin != "direct", r.property_local))
    return rows, notes


@dataclass
class PatternConcept:
    """One blueprint pattern this hub's ontologies are authored against, ready for the
    report's "Core Concepts" sheet: a brief explanation (from the pattern library's own
    ``pattern.md``) paired with one real example drawn from the hub's own properties."""

    pattern_id: str
    title: str
    explanation: str
    example_domain: str
    example_class: str
    example_property: str
    example_excerpt: str


@dataclass
class _PatternExample:
    domain: str
    class_local: str
    property_local: str
    excerpt: str


#: Matches a blueprint-pattern reference inside authored prose, e.g. an ``rdfs:comment`` that
#: reads "as patterns/deferred-relationship requires" or "patterns\\multimodal-order-leg grain
#: 2" (both separators are seen in the wild across this hub's authored comments).
_PATTERN_REF_RE = re.compile(r"patterns[/\\]([a-z][a-z0-9-]+)")


def _pattern_example_excerpt(comment: str, limit: int = 320) -> str:
    """A short, human-readable excerpt of *comment* for the Core Concepts example: its first
    sentence (the plain-English description authors tend to lead with), plus whichever later
    sentence actually names the pattern, if that's a different sentence."""
    normalized = " ".join(comment.split())
    sentences = re.split(r"(?<=[.!?])\s+", normalized)
    if not sentences:
        return normalized[:limit]
    first = sentences[0]
    pattern_sentence = next((s for s in sentences[1:] if "pattern" in s.lower()), "")
    excerpt = f"{first} {pattern_sentence}".strip() if pattern_sentence else first
    if len(excerpt) > limit:
        excerpt = excerpt[: limit - 1].rstrip() + "…"
    return excerpt


def _collect_pattern_examples(ontology_files: list[Path]) -> dict[str, _PatternExample]:
    """Return, for every blueprint pattern referenced in a property's ``rdfs:comment``
    anywhere in *ontology_files*, the first (domain, class, property) that references it --
    a single concrete, real example per pattern, not an invented one.

    Scoped to properties this domain asserts directly (``import_depth == 0``): an imported
    foundation property's own comment isn't this hub's evidence that it exercises the
    pattern. Covers both datatype and object properties (unlike ``_domain_properties``,
    which is scalar-fields-only) because the patterns most worth explaining here --
    ``deferred-relationship``, ``multimodal-order-leg`` -- are usually documented on the
    object property, not a scalar.
    """
    examples: dict[str, _PatternExample] = {}
    for ontology_file in ontology_files:
        domain_name = ontology_file.stem
        loaded = load_ontology(ontology_file, degraded=True, profile=SemanticProfile.RDFS)
        index = loaded.semantic_index
        for record in index.properties:
            if record.provenance.import_depth != 0 or not record.comment:
                continue
            pattern_ids = set(_PATTERN_REF_RE.findall(record.comment))
            if not pattern_ids:
                continue
            class_local = extract_local_name(record.domains[0].uri) if record.domains else ""
            excerpt = _pattern_example_excerpt(record.comment)
            for pattern_id in pattern_ids:
                if pattern_id in examples:
                    continue
                examples[pattern_id] = _PatternExample(
                    domain=domain_name,
                    class_local=class_local,
                    property_local=record.name,
                    excerpt=excerpt,
                )
    return examples


def _discover_patterns_root(hub_root: Path) -> Path | None:
    """Locate the blueprint pattern library's ``patterns/`` directory, checked out either
    inside the hub (``<hub>/ontology-reference-models/...``) or, as this toolkit's own
    dogfood hubs lay it out, as a sibling of the hub root."""
    for candidate in (
        hub_root / "ontology-reference-models" / "blueprints" / "patterns",
        hub_root.parent / "ontology-reference-models" / "blueprints" / "patterns",
    ):
        if candidate.is_dir():
            return candidate
    return None


def _extract_pattern_summary(markdown_text: str, limit: int = 600) -> tuple[str, str]:
    """Pull a pattern's title (its ``# `` heading) and a brief explanation (its own
    ``## Problem`` section's first paragraph) straight out of the pattern library's
    ``pattern.md`` -- authored prose, not a paraphrase, so it stays accurate as patterns
    evolve."""
    lines = markdown_text.splitlines()
    title = ""
    for line in lines:
        if line.startswith("# "):
            title = line[2:].strip()
            break
    problem_lines: list[str] = []
    in_problem = False
    for line in lines:
        if line.startswith("## "):
            if in_problem:
                break
            in_problem = line.strip().lower() == "## problem"
            continue
        if in_problem:
            if not line.strip() and problem_lines:
                break
            if line.strip():
                problem_lines.append(line.strip())
    explanation = " ".join(problem_lines)
    if len(explanation) > limit:
        explanation = explanation[: limit - 1].rstrip() + "…"
    return title, explanation


def _build_core_concepts(
    ontology_files: list[Path], hub_root: Path, notes: list[str]
) -> list[PatternConcept]:
    """Build the "Core Concepts" sheet's content: one entry per blueprint pattern this hub's
    ontologies actually reference, each paired with a real example. Degrades gracefully (an
    empty list, plus an explanatory note) rather than failing the whole report when no
    pattern library is checked out, or when a referenced pattern has no ``pattern.md``.
    """
    examples = _collect_pattern_examples(ontology_files)
    if not examples:
        return []
    patterns_root = _discover_patterns_root(hub_root)
    if patterns_root is None:
        notes.append(
            "Blueprint pattern library not found under 'ontology-reference-models/blueprints/"
            "patterns' (checked the hub root and its parent); skipping the Core Concepts sheet."
        )
        return []
    concepts: list[PatternConcept] = []
    for pattern_id in sorted(examples):
        pattern_md = patterns_root / pattern_id / "pattern.md"
        if not pattern_md.is_file():
            notes.append(f"No pattern.md found for referenced pattern '{pattern_id}'; skipped.")
            continue
        title, explanation = _extract_pattern_summary(pattern_md.read_text(encoding="utf-8"))
        example = examples[pattern_id]
        concepts.append(
            PatternConcept(
                pattern_id=pattern_id,
                title=title or pattern_id,
                explanation=explanation,
                example_domain=example.domain,
                example_class=example.class_local,
                example_property=example.property_local,
                example_excerpt=example.excerpt,
            )
        )
    return concepts


def _fact_source_column_uris(fact: ColumnMappingFact) -> tuple[str, ...]:
    """Return every leaf source-column URI a mapping fact actually reads from.

    A plain ``fields: expression: <Column>`` direct reference leaves ``fact.expression``
    ``None`` (``core/compiler/adapter.py``'s ``_build_field_mappings``), and
    ``fact.source_column_uri`` is the real column -- the only case that may fall back to it.

    Whenever ``fact.expression`` is set, trust ``expression_input_uris(fact.expression)``
    as-is, including an empty result: a literal/null expression (or a CASE whose branches
    are all literals) genuinely has zero source-column leaves, but
    ``adapter.py``'s ``_build_field_mappings`` still needs *some* non-empty
    ``source_column_uri`` for its own bookkeeping and falls back to an ARBITRARY column of
    the relation (``sorted(symbol.uri for symbol in symbols.values())[0]``) in that case --
    falling back to it here too would misreport a constant-valued field as sourced from,
    and sampled from, an unrelated column it never reads.
    """
    if fact.expression is None:
        return (fact.source_column_uri,) if fact.source_column_uri else ()
    return expression_input_uris(fact.expression)


def _source_column_display(
    source_column_uri: str,
    source_columns: dict[str, SourceColumnSample],
    by_table_and_name: dict[tuple[str, str], SourceColumnSample],
) -> tuple[str, str]:
    """Return ``(display_name, sample_value)`` for one leaf source-column URI.

    Prefers the resolved ``SourceColumnSample`` (real table/column name plus a sample
    value, when source vocabulary/sample data exists for it -- issue #298 documents that
    it may not). Falls back to parsing the synthesized ``table_uri/column_name`` key
    ``ResolvedRelation.column_uri()`` (``core/compiler/adapter.py``) always produces for a
    v5 binding, so a column still shows *something* mapped rather than nothing when no
    source vocabulary/sample exists for it.
    """
    column = _resolve_source_column(source_column_uri, source_columns, by_table_and_name)
    if column is not None:
        sample = column.samples[0] if column.samples else ""
        return f"{column.table_name}.{column.name}", sample
    table_uri, sep, name = source_column_uri.rpartition("/")
    if not sep:
        return source_column_uri, ""
    return f"{extract_local_name(table_uri)}.{name}", ""


def run_field_mapping_report(
    *,
    ontologies_path: Path,
    bindings_dir: Path,
    sources_dir: Path,
    hub_root: Path,
    source_system: str,
    domains: tuple[str, ...] = (),
) -> FieldMappingReport:
    """Build a field-mapping report for *source_system* across domain ontologies.

    ``domains``, if given, restricts which domain ontology files are included (matched
    against the ontology file's stem); empty means every domain ontology under
    ``ontologies_path``.
    """
    report = FieldMappingReport(
        source_system=source_system,
        generated_at=datetime.now(timezone.utc).isoformat(),
        ontologies_path=str(ontologies_path),
        bindings_dir=str(bindings_dir),
        sources_dir=str(sources_dir),
    )

    ontology_files = [
        path
        for path in sorted(ontologies_path.glob("**/*.ttl"))
        if is_domain_ontology_stem(path.stem) and (not domains or path.stem in domains)
    ]
    if not ontology_files:
        report.notes.append(f"No domain ontology files found under '{ontologies_path}'.")
        return report

    report.core_concepts = _build_core_concepts(ontology_files, hub_root, report.notes)

    source_columns = load_source_samples(sources_dir)
    by_table_and_name: dict[tuple[str, str], SourceColumnSample] = {
        (column.table_uri, column.name): column for column in source_columns.values()
    }

    # Filtered to *source_system* at the binding-declaration level (the relation's own
    # first dot-segment, e.g. "cargowise" in "cargowise.GlbStaff.sample") rather than via
    # the resolved SourceColumnSample.system: the relation is always present on a
    # resolvable binding, whereas a matching bronze SourceColumn/sample entry may not be
    # (issue #298) -- filtering on that instead would silently drop genuinely-mapped-but
    # -unsampled columns rather than showing them with a blank sample.
    facts, findings = resolve_v5_column_facts(hub_root, bindings_dir, source_system=source_system)
    for finding in findings:
        report.notes.append(finding.message)

    rows_by_property: dict[str, list] = {}
    for fact in facts:
        rows_by_property.setdefault(fact.target_property_uri, []).append(fact)

    for ontology_file in ontology_files:
        domain_name = ontology_file.stem
        domain_properties, domain_notes = _domain_properties(ontology_file)
        report.notes.extend(domain_notes)
        rows: list[FieldMappingRow] = []
        for meta in domain_properties:
            facts = rows_by_property.get(meta.property_uri, [])
            display_columns: list[str] = []
            sample_value = ""
            for fact in facts:
                for source_column_uri in _fact_source_column_uris(fact):
                    display, sample = _source_column_display(
                        source_column_uri, source_columns, by_table_and_name
                    )
                    if display not in display_columns:
                        display_columns.append(display)
                    if not sample_value and sample:
                        sample_value = sample
            rows.append(
                FieldMappingRow(
                    domain=domain_name,
                    class_local=meta.class_local,
                    property_local=meta.property_local,
                    property_uri=meta.property_uri,
                    description=meta.description,
                    range_display=meta.range_display,
                    origin=meta.origin,
                    source_columns=tuple(display_columns),
                    sample_value=sample_value,
                )
            )
        report.rows_by_domain[domain_name] = rows

    return report


#: Shown in "Source Field(s)" for a field with no binding under the selected source system,
#: instead of leaving the cell blank (indistinguishable from a mapped-but-unsampled field).
NO_MAPPING_FOUND = "NO-MAPPING-FOUND"

_HEADER = (
    "Ontology Class (Label)",
    "Ontology Field (Label)",
    "Origin",
    "Ontology Description",
    "Range",
    "Source Field(s)",
    "Source Field Example",
    "Ontology Reference IRI",
)
_WRAP_COLS = {4, 6}  # Ontology Description, Source Field(s) (1-indexed columns)
_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 10


def _row_cells(row: FieldMappingRow) -> tuple[str, str, str, str, str, str, str, str]:
    return (
        row.class_local,
        row.property_local,
        row.origin,
        row.description,
        row.range_display,
        "; ".join(row.source_columns) if row.source_columns else NO_MAPPING_FOUND,
        row.sample_value,
        row.property_uri,
    )


def _style_header_row(sheet) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _autosize_columns(sheet) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, width in widths.items():
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = min(
            max(width + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH
        )
        if col in _WRAP_COLS:
            for cell in sheet[letter][1:]:  # data rows only, header already aligned
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _set_wrapped_row_height(sheet, row_idx: int, text: str, width_chars: int = 120) -> None:
    lines = max(1, -(-len(text) // width_chars))  # ceil division, no non-str-length import
    sheet.row_dimensions[row_idx].height = 15 * lines


def _write_core_concepts_sheet(sheet, concepts: list[PatternConcept]) -> None:
    from openpyxl.styles import Alignment, Font

    sheet.append(["Core Concepts"])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append(
        [
            "The blueprint patterns this hub's ontologies are authored against, each with one "
            "real example drawn from this hub's own classes and properties -- not a generic "
            "illustration."
        ]
    )
    _set_wrapped_row_height(sheet, 2, sheet["A2"].value)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.append([])

    for concept in concepts:
        sheet.append([concept.title])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=12)

        sheet.append([concept.explanation])
        row = sheet.max_row
        sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        _set_wrapped_row_height(sheet, row, concept.explanation)

        sheet.append(["Example (from this hub):"])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, italic=True)

        example_line = (
            f"{concept.example_domain}: :{concept.example_class} -> "
            f":{concept.example_property} -- {concept.example_excerpt}"
        )
        sheet.append([example_line])
        row = sheet.max_row
        sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        _set_wrapped_row_height(sheet, row, example_line)

        sheet.append([])

    sheet.column_dimensions["A"].width = 120


def write_field_mapping_workbook(report: FieldMappingReport, output_path: Path) -> None:
    """Write *report* to a styled .xlsx workbook, one worksheet per domain plus a cover sheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to write the field-mapping report. "
            "Install it with: pip install kairos-ontology-toolkit[flatfile]"
        ) from exc

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Overview"
    cover.append(["Field mapping report"])
    cover["A1"].font = Font(bold=True, size=14)
    cover.append(["Source system", report.source_system])
    cover.append(["Generated at", report.generated_at])
    cover.append(["Ontologies", report.ontologies_path])
    cover.append(["Bindings", report.bindings_dir])
    cover.append(["Sources", report.sources_dir])
    cover.append([])
    cover.append(["Scope: owl:DatatypeProperty fields only. Object properties / relationship"])
    cover.append(["joins are not yet included -- see issue tracker."])
    if report.notes:
        cover.append([])
        cover.append(["Notes"])
        for note in report.notes:
            cover.append([note])
    for row in cover.iter_rows(min_row=2, max_row=6, max_col=1):
        row[0].font = Font(bold=True)
    cover.column_dimensions["A"].width = 16
    cover.column_dimensions["B"].width = 80

    if report.core_concepts:
        concepts_sheet = workbook.create_sheet(title="Core Concepts", index=1)
        _write_core_concepts_sheet(concepts_sheet, report.core_concepts)

    for domain_name, rows in report.rows_by_domain.items():
        sheet_title = domain_name[:31]
        sheet = workbook.create_sheet(title=sheet_title)
        sheet.append(_HEADER)
        for row in rows:
            sheet.append(_row_cells(row))
        _style_header_row(sheet)
        _autosize_columns(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
