# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Cnext.eu
"""Import-Flatfile — create source schema YAML from CSV/Excel/Parquet flat files.

Reads CSV, Excel, or Parquet files and produces the same intermediate YAML + samples
format that extract-schema generates from live databases. The output directory can then
be passed to ``import-source`` to generate bronze vocabulary TTL.

Pipeline: CSV/XLSX/Parquet → _manifest.yaml + {table}.yaml + {table}.samples.yaml
          → import-source → .vocabulary.ttl → analyse-sources
"""

from __future__ import annotations

import csv
import copy
import importlib
import logging
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml

from ._samples import (
    SAMPLE_PRIVACY_POLICY,
    SAMPLE_PRIVACY_VERSION,
    assert_no_unredacted_sample_pii,
    redact_sample_rows,
)

# Increase CSV field size limit to handle large fields (e.g., Oracle exports).
# On Windows 64-bit, sys.maxsize exceeds C long max, so cap at 2^31 - 1.
csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

logger = logging.getLogger(__name__)

# Maximum rows to read for type inference (to avoid loading huge files into memory)
DEFAULT_MAX_ROWS = 1000
DEFAULT_SAMPLE_SIZE = 5

#: Distinct values captured per column (DD-166).
#:
#: Deliberately separate from DEFAULT_SAMPLE_SIZE, which the two once shared. They are
#: different artifacts with different risk: a sample *row* correlates every column of
#: one real record and is the PII-sensitive output, while distinct column *values* are
#: de-correlated and are what the alignment step reasons over to tell a governed code
#: list from free text. Five was too few for that judgement; twenty covers most code
#: lists. Raising the row count to match would have widened the PII surface for no
#: modelling gain, which is why these are now two numbers.
DEFAULT_DISTINCT_VALUES = 20

# File extensions directory mode will attempt to read. Anything else in the
# directory is ignored silently (not counted as a failure). Exported so callers
# can report "M of K file(s)" against the same candidate set the loop uses.
SUPPORTED_FLATFILE_SUFFIXES = frozenset({".csv", ".xlsx", ".xls", ".parquet"})

# Known lakehouse/ingestion metadata columns that are typically technical noise.
# Columns matching these names (case-insensitive) and appearing in all tables with
# distinctCount=1 are auto-excluded unless --keep-technical is set.
KNOWN_TECHNICAL_COLUMNS = frozenset(
    {
        "volume",
        "subfolder",
        "table",
        "last_ingest_date",
        "rowversion",
    }
)

# --------------------------------------------------------------------------- #
# Type Inference
# --------------------------------------------------------------------------- #

# Date/datetime patterns
_DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),  # 2024-01-15
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),  # 15/01/2024
    re.compile(r"^\d{2}-\d{2}-\d{4}$"),  # 15-01-2024
]
_DATETIME_PATTERNS = [
    # End-anchored like _DATE_PATTERNS above: without the ``$`` any free-text column
    # whose values merely BEGIN with a timestamp (audit trails, comment logs, contact
    # notes) was typed ``datetime`` (#302).
    re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(?::\d{2}(?:\.\d+)?)?(?:Z|[+-]\d{2}:\d{2})?$"),
]
_BOOL_VALUES = {"true", "false", "1", "0", "yes", "no", "y", "n"}


def infer_column_type(values: list[str]) -> str:
    """Infer SQL-like data type from a list of string values.

    Args:
        values: Non-empty string values from the column (nulls/blanks excluded).

    Returns:
        One of: 'bigint', 'int', 'decimal', 'date', 'datetime', 'bit', 'varchar(max)'
    """
    if not values:
        return "varchar(max)"

    # Check for boolean
    if all(v.lower().strip() in _BOOL_VALUES for v in values):
        return "bit"

    # Check for datetime (before date — datetime is a superset)
    if all(any(p.match(v.strip()) for p in _DATETIME_PATTERNS) for v in values):
        return "datetime"

    # Check for date
    if all(any(p.match(v.strip()) for p in _DATE_PATTERNS) for v in values):
        return "date"

    # Check for integer
    int_count = 0
    for v in values:
        stripped = v.strip()
        try:
            int(stripped)
            int_count += 1
        except ValueError:
            break
    if int_count == len(values):
        max_val = max(abs(int(v.strip())) for v in values)
        return "bigint" if max_val > 2_147_483_647 else "int"

    # Check for decimal/float
    float_count = 0
    for v in values:
        stripped = v.strip()
        try:
            float(stripped)
            float_count += 1
        except ValueError:
            break
    if float_count == len(values):
        return "decimal"

    return "varchar(max)"


# --------------------------------------------------------------------------- #
# CSV Reading
# --------------------------------------------------------------------------- #


def read_csv_table(
    path: Path,
    max_rows: int = DEFAULT_MAX_ROWS,
    sample_size: int = DEFAULT_SAMPLE_SIZE,
    table_stem: str | None = None,
) -> dict[str, Any]:
    """Read a CSV file and produce a table data dict.

    Args:
        path: Path to the .csv file.
        max_rows: Maximum rows to read for type inference.
        sample_size: Number of sample rows to store.
        table_stem: Override for the table name (default: ``path.stem``). Used by
            ``--recursive`` directory mode to derive a collision-safe name from
            the file's path relative to ``--from`` instead of the bare filename.

    Returns:
        Dict with keys: name, row_count, rows_sampled, columns, sample_rows.
        ``row_count`` is the TRUE table cardinality and is ``None`` when the
        read was capped at ``max_rows`` (the file has more rows than were
        profiled); ``rows_sampled`` is always the number of rows actually read.
    """
    table_name = table_stem if table_stem is not None else path.stem

    with open(path, encoding="utf-8-sig", newline="") as f:
        # Sniff the dialect
        sample_text = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample_text, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(f, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError(f"No headers found in CSV file: {path}")

        headers = list(reader.fieldnames)
        all_rows: list[dict[str, str]] = []
        # Cap-hit detection (#422): the break can only fire when a row BEYOND
        # the cap exists — natural loop exhaustion (even at exactly max_rows
        # rows) means the file was read in full and len(all_rows) is the true
        # table cardinality.
        capped = False
        for i, row in enumerate(reader):
            if i >= max_rows:
                capped = True
                break
            all_rows.append(row)

    rows_sampled = len(all_rows)
    row_count = None if capped else rows_sampled

    # Build column metadata
    columns = []
    for pos, col_name in enumerate(headers, start=1):
        non_empty_values = [
            (row.get(col_name) or "").strip()
            for row in all_rows
            if (row.get(col_name) or "").strip()
        ]
        distinct_values = list(dict.fromkeys(non_empty_values))
        nullable = len(non_empty_values) < rows_sampled

        col_dict: dict[str, Any] = {
            "name": col_name,
            "data_type": infer_column_type(distinct_values[:100]),
            "ordinal_position": pos,
            "nullable": nullable,
        }
        if distinct_values:
            col_dict["distinct_count"] = len(distinct_values)
            col_dict["samples"] = distinct_values[:DEFAULT_DISTINCT_VALUES]

        columns.append(col_dict)

    # Sample rows (raw dicts for .samples.yaml)
    sample_rows = [{k: v for k, v in row.items() if v} for row in all_rows[:sample_size]]

    return {
        "name": table_name,
        "row_count": row_count,
        "rows_sampled": rows_sampled,
        "columns": columns,
        "sample_rows": sample_rows,
    }


# --------------------------------------------------------------------------- #
# Excel Reading
# --------------------------------------------------------------------------- #


def read_xlsx_tables(
    path: Path,
    max_rows: int = DEFAULT_MAX_ROWS,
    sample_size: int = DEFAULT_SAMPLE_SIZE,
    table_stem: str | None = None,
) -> list[dict[str, Any]]:
    """Read an Excel workbook and produce one table dict per worksheet.

    Requires openpyxl (install via: pip install kairos-ontology-toolkit[flatfile]).
    Legacy ``.xls`` is never routed here — callers must reject it before calling
    (see the dispatch sites in ``run_import_flatfile``): openpyxl raises
    ``InvalidFileException`` for ``.xls``, which is neither ``ValueError`` nor
    ``ImportError``.

    Args:
        path: Path to the .xlsx file.
        max_rows: Maximum rows to read per sheet for type inference.
        sample_size: Number of sample rows to store.
        table_stem: Override for the file-level name component (default:
            ``path.stem``). Used by ``--recursive`` directory mode to derive a
            collision-safe name from the file's path relative to ``--from``.

    Returns:
        List of table data dicts (same format as read_csv_table).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel support. "
            "Install with: pip install kairos-ontology-toolkit[flatfile]"
        )

    stem = table_stem if table_stem is not None else path.stem
    wb = load_workbook(path, read_only=True, data_only=True)
    tables = []

    has_multiple_sheets = len(wb.sheetnames) > 1
    for sheet_name in wb.sheetnames:
        table_name = f"{stem}__{sheet_name}" if has_multiple_sheets else stem
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # First row = headers
        try:
            headers_row = next(rows_iter)
        except StopIteration:
            continue

        headers = [str(h).strip() if h else f"Column{i}" for i, h in enumerate(headers_row, 1)]
        if not any(h for h in headers):
            continue

        # Read data rows
        all_rows: list[dict[str, str]] = []
        # Cap-hit detection (#422), per sheet: the break fires only when a row
        # BEYOND the cap exists; natural exhaustion = full read = true count.
        # Note: openpyxl can yield "ghost" rows (formatting-only, all-None) past
        # the real data, so a sheet with exactly max_rows real rows followed by
        # ghost rows is treated as capped and its row_count conservatively
        # omitted. That under-claims knowledge but never asserts a false count.
        capped = False
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                capped = True
                break
            row_dict = {}
            for col_name, val in zip(headers, row):
                row_dict[col_name] = str(val) if val is not None else ""
            all_rows.append(row_dict)

        rows_sampled = len(all_rows)
        if rows_sampled == 0:
            continue
        row_count = None if capped else rows_sampled

        # Build column metadata
        columns = []
        for pos, col_name in enumerate(headers, start=1):
            non_empty_values = [
                (row.get(col_name) or "").strip()
                for row in all_rows
                if (row.get(col_name) or "").strip()
            ]
            distinct_values = list(dict.fromkeys(non_empty_values))
            nullable = len(non_empty_values) < rows_sampled

            col_dict: dict[str, Any] = {
                "name": col_name,
                "data_type": infer_column_type(distinct_values[:100]),
                "ordinal_position": pos,
                "nullable": nullable,
            }
            if distinct_values:
                col_dict["distinct_count"] = len(distinct_values)
                col_dict["samples"] = distinct_values[:DEFAULT_DISTINCT_VALUES]

            columns.append(col_dict)

        sample_rows = [{k: v for k, v in row.items() if v} for row in all_rows[:sample_size]]

        tables.append(
            {
                "name": table_name,
                "row_count": row_count,
                "rows_sampled": rows_sampled,
                "columns": columns,
                "sample_rows": sample_rows,
            }
        )

    wb.close()
    return tables


# --------------------------------------------------------------------------- #
# Parquet Reading
# --------------------------------------------------------------------------- #


def _arrow_type_to_sql(arrow_type: Any) -> str:
    """Map a pyarrow data type to the SQL-like type vocabulary.

    Parquet carries a reliable typed schema, so types are mapped directly
    rather than inferred from string values.

    Args:
        arrow_type: A ``pyarrow.DataType`` instance.

    Returns:
        One of: bigint, int, decimal, date, datetime, bit, varchar(max).
    """
    import pyarrow as pa

    if pa.types.is_boolean(arrow_type):
        return "bit"
    if pa.types.is_int64(arrow_type) or pa.types.is_uint64(arrow_type):
        return "bigint"
    if pa.types.is_integer(arrow_type):
        # int8/16/32 and uint8/16/32
        if pa.types.is_uint32(arrow_type):
            return "bigint"
        return "int"
    if pa.types.is_floating(arrow_type) or pa.types.is_decimal(arrow_type):
        return "decimal"
    if pa.types.is_timestamp(arrow_type):
        return "datetime"
    if pa.types.is_date(arrow_type):
        return "date"
    return "varchar(max)"


def _arrow_column_to_pylist(column: Any, field_type: Any) -> list:
    """Materialise one Arrow column as Python values without needing a tz database.

    tz-aware timestamps are normalised to UTC and rendered as RFC-3339 text
    (``2024-01-15 10:30:00+00:00``) so the offset is explicit and the value
    matches ``core/_samples.py::_ISO_DATE_OR_DATETIME_RE``. Casting to a
    tz-naive timestamp instead would silently drop the offset and, for a
    non-UTC source zone, shift the displayed wall clock.

    tz-NAIVE timestamp columns are left to ``to_pylist()`` as-is — they carry
    no offset, and rendering them without one is honest about what the source
    actually recorded. Only tz-aware columns go through the UTC-normalising
    path above; the difference in rendering between the two is intentional.

    Any Arrow type this cannot materialise (e.g. an exotic type pyarrow itself
    fails to convert) degrades to a column of ``None`` values with a logged
    warning, rather than aborting the whole import.
    """
    import pyarrow as pa

    try:
        if pa.types.is_timestamp(field_type) and field_type.tz is not None:
            naive_values = column.cast(pa.timestamp(field_type.unit)).to_pylist()
            return [
                v.replace(tzinfo=timezone.utc).isoformat(sep=" ") if v is not None else None
                for v in naive_values
            ]
        return column.to_pylist()
    except Exception as exc:
        logger.warning(
            "Column of type %s could not be materialised: %s",
            field_type,
            exc,
        )
        return [None] * len(column)


def read_parquet_table(
    path: Path,
    max_rows: int = DEFAULT_MAX_ROWS,
    sample_size: int = DEFAULT_SAMPLE_SIZE,
    table_stem: str | None = None,
) -> dict[str, Any]:
    """Read a single Parquet file into a table data dict.

    Only sample data is read — at most ``max_rows`` rows are pulled (a single
    Arrow batch). The full Parquet body is never loaded into memory, mirroring
    the CSV/Excel readers. Column data types come directly from the Parquet
    schema; sample values are stringified to match the YAML output format.

    Args:
        path: Path to the .parquet file.
        max_rows: Maximum rows to read for sampling.
        sample_size: Number of sample rows to store.
        table_stem: Override for the table name (default: ``path.stem``). Used by
            ``--recursive`` directory mode to derive a collision-safe name from
            the file's path relative to ``--from`` instead of the bare filename.

    Returns:
        Dict with keys: name, row_count, rows_sampled, columns, sample_rows.
        ``row_count`` is the TRUE table cardinality (from the Parquet footer
        metadata, covering all row groups); ``rows_sampled`` is the size of the
        profiling window actually read (#422).
    """
    try:
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError(
            "pyarrow is required for Parquet support. "
            "Install with: pip install kairos-ontology-toolkit[parquet]"
        )

    table_name = table_stem if table_stem is not None else path.stem
    pf = pq.ParquetFile(path)

    # True table cardinality comes from the Parquet footer metadata (#422) —
    # free to read and covering ALL row groups. (Previously the first batch's
    # size doubled as row_count, which both conflated window size with
    # cardinality and undercounted multi-row-group files.)
    row_count = pf.metadata.num_rows

    # Read at most max_rows rows (a single batch) — never the whole file.
    batch = None
    for b in pf.iter_batches(batch_size=max_rows):
        batch = b
        break

    schema = pf.schema_arrow
    headers = list(schema.names)

    if batch is None:
        # Empty parquet file: still emit columns from the schema.
        columns = [
            {
                "name": name,
                "data_type": _arrow_type_to_sql(schema.field(name).type),
                "ordinal_position": pos,
                "nullable": True,
            }
            for pos, name in enumerate(headers, start=1)
        ]
        return {
            "name": table_name,
            "row_count": row_count,
            "rows_sampled": 0,
            "columns": columns,
            "sample_rows": [],
        }

    # Size of the profiling window. Everything derived from the read rows
    # (nullable, distinct/sample slicing) must use this, NOT row_count — the
    # window may be a strict subset of the table.
    rows_sampled = batch.num_rows

    # Build per-row string dicts for samples (mirrors CSV/XLSX format).
    # Naive timestamp[us] columns pass straight through to_pylist(); tz-aware
    # columns route through _arrow_column_to_pylist() to avoid pyarrow's
    # zoneinfo lookup (no tz database is required on Windows). See that
    # helper's docstring for why the two are rendered differently.
    column_values: dict[str, list] = {
        name: _arrow_column_to_pylist(batch.column(i), schema.field(name).type)
        for i, name in enumerate(headers)
    }

    columns = []
    for pos, col_name in enumerate(headers, start=1):
        raw_values = column_values[col_name]
        non_empty_values = [str(v).strip() for v in raw_values if v is not None and str(v).strip()]
        distinct_values = list(dict.fromkeys(non_empty_values))
        nullable = len(non_empty_values) < rows_sampled

        col_dict: dict[str, Any] = {
            "name": col_name,
            "data_type": _arrow_type_to_sql(schema.field(col_name).type),
            "ordinal_position": pos,
            "nullable": nullable,
        }
        if distinct_values:
            col_dict["distinct_count"] = len(distinct_values)
            col_dict["samples"] = distinct_values[:DEFAULT_DISTINCT_VALUES]

        columns.append(col_dict)

    # Sample rows (raw dicts for .samples.yaml), stringified, empties dropped.
    sample_rows = []
    for r in range(min(sample_size, rows_sampled)):
        row = {
            col_name: str(column_values[col_name][r]).strip()
            for col_name in headers
            if column_values[col_name][r] is not None and str(column_values[col_name][r]).strip()
        }
        sample_rows.append(row)

    return {
        "name": table_name,
        "row_count": row_count,
        "rows_sampled": rows_sampled,
        "columns": columns,
        "sample_rows": sample_rows,
    }


# --------------------------------------------------------------------------- #
# Output Writing
# --------------------------------------------------------------------------- #


def write_source_dir(
    tables: list[dict[str, Any]],
    system_name: str,
    output_dir: Path,
    platform: str = "flatfile",
    redact_pii: bool = False,
) -> Path:
    """Write table data to the standard source directory format.

    Creates:
      - _manifest.yaml
      - {table}.yaml per table (schema metadata)
      - {table}.samples.yaml per table (sample rows)

    Args:
        tables: List of table data dicts (from read_csv_table or read_xlsx_tables).
        system_name: System name for the source.
        output_dir: Target directory for output files.
        platform: Platform identifier (default: "flatfile").

    Returns:
        Path to the output directory.
    """
    table_names = [str(table["name"]) for table in tables]
    duplicate_names = sorted({name for name in table_names if table_names.count(name) > 1})
    if duplicate_names:
        names = ", ".join(duplicate_names)
        raise ValueError(f"Duplicate final table name(s) in flatfile import: {names}")

    # Sanitize and validate every sample before publishing any artifact.
    safe_tables = copy.deepcopy(tables)
    # Redaction and its assertions are gated as one unit (issue #692).
    # ``assert_no_unredacted_sample_pii`` is a *post-condition* of ``redact_sample_rows``, not
    # an independent gate -- skipping the redaction while keeping the assert would raise
    # ``SamplePrivacyError`` on the raw values and kill the import outright.
    for table in safe_tables if redact_pii else []:
        column_types = {
            str(col.get("name", "")): str(col.get("data_type", "unknown"))
            for col in table.get("columns", [])
        }
        safe_rows, _ = redact_sample_rows(
            table.get("sample_rows", []),
            table=str(table["name"]),
            column_types=column_types,
        )
        assert_no_unredacted_sample_pii(safe_rows, table=str(table["name"]))
        table["sample_rows"] = safe_rows

        # Per-column distinct values were previously left raw and then dropped on write
        # -- dropping them *was* the privacy control. They are now published (the
        # alignment step reasons over them), so they must clear the same bar as rows.
        # Reusing redact_sample_rows rather than writing a second detector keeps one
        # audited implementation: each value becomes a one-cell row for that column.
        for col in table.get("columns", []):
            values = col.get("samples") or []
            if not values:
                continue
            name = str(col.get("name", ""))
            pseudo_rows = [{name: value} for value in values]
            safe_values, _ = redact_sample_rows(
                pseudo_rows, table=str(table["name"]), column_types=column_types
            )
            assert_no_unredacted_sample_pii(safe_values, table=str(table["name"]))
            # Redaction collapses distinct values to identical tokens; dedupe again so a
            # PII column contributes one masked token, not twenty copies of it.
            col["samples"] = list(
                dict.fromkeys(
                    row[name] for row in safe_values if row.get(name) not in (None, "")
                )
            )

    output_dir.mkdir(parents=True, exist_ok=True)

    # Write manifest
    manifest = {
        "version": "1.2",
        "system": system_name,
        "platform": platform,
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "connection": {},
        "tables": [t["name"] for t in safe_tables],
        # State the policy actually applied. Stamping ``redact-detected-pii`` on artifacts
        # written without it would be the same overstatement DD-075's first amendment exists
        # to prevent -- and the version is omitted with it, since there is no policy version
        # to record when no policy ran.
        "sample_privacy": (
            {"policy": SAMPLE_PRIVACY_POLICY, "version": SAMPLE_PRIVACY_VERSION}
            if redact_pii
            else {"policy": "none"}
        ),
    }
    with open(output_dir / "_manifest.yaml", "w", encoding="utf-8") as f:
        yaml.dump(manifest, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

    # Write per-table YAML + samples
    for table in safe_tables:
        tbl_name = table["name"]

        # Schema metadata (without sample_rows — those go in .samples.yaml)
        table_yaml: dict[str, Any] = {"name": tbl_name}
        # v1.2 semantics (#422): row_count is TRUE table cardinality and is
        # OMITTED when unknown (capped read); rows_sampled is the profiling
        # window size and is always present for flatfile-read tables.
        if table.get("row_count") is not None:
            table_yaml["row_count"] = table["row_count"]
        if table.get("rows_sampled") is not None:
            table_yaml["rows_sampled"] = table["rows_sampled"]
        # ``samples`` (sanitized distinct column values) is published; ``sample_rows``
        # still lives only in <table>.samples.yaml. Column values are de-correlated --
        # they cannot be reassembled into a record the way a row can.
        table_yaml["columns"] = list(table["columns"])
        with open(output_dir / f"{tbl_name}.yaml", "w", encoding="utf-8") as f:
            yaml.dump(table_yaml, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

        # Sample rows
        sample_rows = table.get("sample_rows", [])
        if sample_rows:
            samples_data = {
                "extracted_at": manifest["extracted_at"],
                "table": tbl_name,
                "schema": "",
                "sample_privacy": manifest["sample_privacy"],
                "rows": sample_rows,
            }
            with open(output_dir / f"{tbl_name}.samples.yaml", "w", encoding="utf-8") as f:
                yaml.dump(
                    samples_data, f, default_flow_style=False, sort_keys=False, allow_unicode=True
                )
        else:
            stale_samples = output_dir / f"{tbl_name}.samples.yaml"
            if stale_samples.exists():
                stale_samples.unlink()

    logger.info("Written %d table(s) to %s", len(tables), output_dir)
    return output_dir


def detect_technical_columns(tables: list[dict[str, Any]]) -> set[str]:
    """Detect columns that are likely technical/metadata noise.

    A column is flagged as technical if it appears in ALL tables with
    distinctCount=1 and its name (case-insensitive) matches a known
    lakehouse metadata pattern.

    ``tables`` here is whatever ``run_import_flatfile`` successfully read —
    the "ALL tables" gate (``count == num_tables``) is relative to that
    successfully-read set, not the full set of files present on disk. If
    directory mode skipped some unreadable files, auto-exclusion runs against
    the surviving tables only, so which columns get auto-excluded can vary
    run to run as the set of readable files changes.

    Args:
        tables: List of table data dicts.

    Returns:
        Set of column names to exclude.
    """
    if not tables:
        return set()

    # Find columns present in every table with distinctCount=1
    candidates: dict[str, int] = {}
    for tbl in tables:
        for col in tbl.get("columns", []):
            name = col["name"]
            if col.get("distinct_count", 0) == 1:
                candidates[name] = candidates.get(name, 0) + 1

    num_tables = len(tables)
    technical = set()
    for name, count in candidates.items():
        if count == num_tables and name.lower() in KNOWN_TECHNICAL_COLUMNS:
            technical.add(name)

    return technical


def exclude_columns_from_tables(
    tables: list[dict[str, Any]], columns_to_exclude: set[str]
) -> list[dict[str, Any]]:
    """Remove specified columns from all tables.

    Args:
        tables: List of table data dicts (modified in place and returned).
        columns_to_exclude: Set of column names to remove (case-insensitive).

    Returns:
        The modified tables list.
    """
    if not columns_to_exclude:
        return tables

    exclude_lower = {c.lower() for c in columns_to_exclude}
    for tbl in tables:
        tbl["columns"] = [col for col in tbl["columns"] if col["name"].lower() not in exclude_lower]
        # Also strip excluded columns from sample rows
        tbl["sample_rows"] = [
            {k: v for k, v in row.items() if k.lower() not in exclude_lower}
            for row in tbl.get("sample_rows", [])
        ]
    return tables


# --------------------------------------------------------------------------- #
# Main Orchestration
# --------------------------------------------------------------------------- #


def list_flatfile_candidates(source_dir: Path, recursive: bool = False) -> list[Path]:
    """Return the sorted files in ``source_dir`` directory mode will try to read.

    Callers reporting "M of K file(s) could not be read" must derive K from this
    same helper, so the denominator can never drift from the set the import loop
    actually attempts.

    Directory mode is non-recursive by default: only files directly inside
    ``source_dir`` are candidates (``Path.iterdir()``), matching a single flat
    export folder. Pass ``recursive=True`` to walk the full subtree
    (``Path.rglob("*")``) for a nested export tree — callers doing this must
    also derive table names from each file's path relative to ``source_dir``
    (see ``run_import_flatfile``), or same-basename files in different
    subdirectories will collide.
    """
    if recursive:
        return sorted(
            f
            for f in source_dir.rglob("*")
            if f.is_file() and f.suffix.lower() in SUPPORTED_FLATFILE_SUFFIXES
        )
    return [
        f for f in sorted(source_dir.iterdir()) if f.suffix.lower() in SUPPORTED_FLATFILE_SUFFIXES
    ]


def _relative_table_name(path: Path, base: Path) -> str:
    """Derive a filesystem-safe table-name stem from ``path`` relative to ``base``.

    Only used in ``--recursive`` directory mode, where nested files can share a
    basename (e.g. ``2024/orders.csv`` vs ``2025/orders.csv``) that would
    otherwise collide once flattened into a single directory of
    ``{table}.yaml`` files. Lowercases and collapses path separators/anything
    non-alphanumeric to a single hyphen, matching the slug style already used
    for nested business-discovery documents
    (``discovery_extraction.slugify_source_name``). The exact-duplicate guard
    in ``write_source_dir`` remains the backstop if two distinct relative
    paths still collapse to the same slug.
    """
    rel = path.relative_to(base).with_suffix("")
    slug = re.sub(r"[^a-z0-9]+", "-", str(rel).replace("\\", "/").lower()).strip("-")
    return slug or path.stem.lower()


# Suffix -> (module name to probe, extra name that provides it). Only suffixes
# whose reader needs an optional dependency are listed. ``.xls`` is deliberately
# absent: it is never actually supported (see the explicit dispatch-site guard
# in ``run_import_flatfile``), so it must not be reported as an "install
# openpyxl and it'll work" problem — installing openpyxl would not fix it.
_EXTRA_REQUIREMENTS: dict[str, tuple[str, str]] = {
    ".xlsx": ("openpyxl", "flatfile"),
    ".parquet": ("pyarrow", "parquet"),
}


def missing_flatfile_extras(candidates: list[Path]) -> dict[str, str]:
    """Return install commands for optional extras ``candidates`` need but lack.

    Keyed on whether the backing module actually imports, not on suffix alone: a
    directory full of ``.xlsx`` files reports nothing here once ``openpyxl`` is
    installed, even though every candidate has that suffix. Only extras actually
    required by at least one candidate's suffix are probed, so a directory with
    no Excel or Parquet files never mentions either.

    This exists so the CLI can preflight a whole batch and fail once, before
    ``run_import_flatfile`` would otherwise raise the same ``ImportError`` once
    per candidate file (directory mode's per-file ``except Exception`` catches
    it independently for each file) — see issue #407. ``run_import_flatfile``
    itself does not call this; it is opt-in for the CLI (and any other library
    caller) so a corrupted/mislabeled single file still gets its own precise
    error instead of being swept into an "extras missing" diagnosis.

    Args:
        candidates: Files that will be attempted (from ``list_flatfile_candidates``
            for directory mode, or a single-element list for one file).

    Returns:
        Dict of extra name -> ``pip install ...`` command, for extras actually
        needed (by suffix) and not importable. Empty when nothing is missing.
    """
    needed: dict[str, str] = {}  # extra name -> module name
    for candidate in candidates:
        requirement = _EXTRA_REQUIREMENTS.get(candidate.suffix.lower())
        if requirement:
            module_name, extra = requirement
            needed[extra] = module_name

    missing: dict[str, str] = {}
    for extra, module_name in needed.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing[extra] = f"pip install kairos-ontology-toolkit[{extra}]"
    return missing


def run_import_flatfile(
    source_path: Path,
    system_name: str | None = None,
    output_dir: Path | None = None,
    max_rows: int = DEFAULT_MAX_ROWS,
    sample_size: int = DEFAULT_SAMPLE_SIZE,
    exclude_columns: set[str] | None = None,
    keep_technical: bool = False,
    return_count: bool = False,
    recursive: bool = False,
    redact_pii: bool = False,
) -> Path | tuple[Path, int, int, list[tuple[str, str]]]:
    """Orchestrate the flatfile import workflow.

    Accepts a single CSV, single XLSX, single Parquet, or a directory containing
    CSV/XLSX/Parquet files.

    Directory mode is non-recursive by default — only files directly inside
    ``source_path`` are read (see ``list_flatfile_candidates``); nested
    subdirectories are ignored unless ``recursive=True``. When recursive, each
    file's table name is derived from its path relative to ``source_path``
    (lowercased, separators collapsed) instead of the bare filename, so
    same-basename files in different subdirectories don't collide once
    flattened into one output directory (the duplicate-name guard in
    ``write_source_dir`` is the backstop if a collision still happens).

    Exit-code policy — directory mode is partial-failure tolerant; single-file
    mode is fail-fast:
    - Directory mode: each file is read independently. A file that raises
      (corrupt/truncated parquet, malformed CSV, missing openpyxl, legacy
      ``.xls``, ...) is skipped and recorded in the returned ``failures`` list;
      the remaining readable files are still imported and written. This
      mirrors the ``catalog-test``/``propose-alignment`` convention of partial
      success over aborting a whole run for one bad input.
    - Zero readable files (directory has no CSV/XLSX/Parquet, or every file
      present failed to read) is a hard failure: ``ValueError`` is raised and
      nothing is written. Callers (the CLI) map this to a non-zero exit.
    - Single-file mode (a path pointing directly at one .csv/.xlsx/.parquet)
      always fails fast — an unsupported extension or a read error propagates
      immediately as the underlying exception. There is no "partial" outcome
      for a single file.
    - Legacy ``.xls`` is a recognized candidate (so directory mode reports it
      by name instead of silently ignoring it) but is never actually readable:
      both dispatch sites below raise ``ValueError`` for it explicitly, rather
      than routing it to the openpyxl-based Excel reader, which raises
      ``InvalidFileException`` — neither a ``ValueError`` nor an
      ``ImportError``, so it would otherwise escape single-file mode as an
      unhandled exception.

    Args:
        source_path: Path to CSV, XLSX file, or directory.
        system_name: Override system name (default: derived from path).
        output_dir: Output directory (default: integration/sources/{system}/).
        max_rows: Maximum rows to read for type inference.
        sample_size: Number of sample rows to store.
        exclude_columns: Explicit set of column names to exclude.
        keep_technical: If True, skip auto-detection of technical columns.
        return_count: If True, return
            ``(output_dir, table_count, sample_file_count, failures)`` where
            ``failures`` is a list of ``(filename, "ExceptionType: message")``
            tuples for files skipped in directory mode (empty otherwise).
        recursive: If True, walk the full subtree of a directory input instead
            of only its top level. No effect on single-file input.

    Returns:
        Path to the output directory, or a tuple with counts from this import.
    """
    tables: list[dict[str, Any]] = []
    failures: list[tuple[str, str]] = []

    if source_path.is_file():
        suffix = source_path.suffix.lower()
        default_name = source_path.stem

        if suffix == ".csv":
            tables.append(read_csv_table(source_path, max_rows, sample_size))
        elif suffix == ".xls":
            raise ValueError("legacy .xls is not supported — convert to .xlsx")
        elif suffix == ".xlsx":
            tables.extend(read_xlsx_tables(source_path, max_rows, sample_size))
        elif suffix == ".parquet":
            tables.append(read_parquet_table(source_path, max_rows, sample_size))
        else:
            raise ValueError(f"Unsupported file type: {suffix}. Use .csv, .xlsx, or .parquet")

    elif source_path.is_dir():
        default_name = source_path.name
        for f in list_flatfile_candidates(source_path, recursive=recursive):
            suffix = f.suffix.lower()
            table_stem = _relative_table_name(f, source_path) if recursive else None
            try:
                if suffix == ".csv":
                    tables.append(read_csv_table(f, max_rows, sample_size, table_stem=table_stem))
                elif suffix == ".xls":
                    raise ValueError("legacy .xls is not supported — convert to .xlsx")
                elif suffix == ".xlsx":
                    tables.extend(read_xlsx_tables(f, max_rows, sample_size, table_stem=table_stem))
                elif suffix == ".parquet":
                    tables.append(
                        read_parquet_table(f, max_rows, sample_size, table_stem=table_stem)
                    )
            except Exception as exc:
                logger.warning("Skipping unreadable file %s: %s", f.name, exc)
                failures.append((f.name, f"{type(exc).__name__}: {exc}"))

        if not tables:
            if failures:
                names = ", ".join(f"{name} ({reason})" for name, reason in failures)
                raise ValueError(
                    f"No CSV, Excel, or Parquet files could be read in: {source_path} "
                    f"(all {len(failures)} file(s) failed: {names})"
                )
            raise ValueError(f"No CSV, Excel, or Parquet files found in: {source_path}")
    else:
        raise ValueError(f"Path does not exist: {source_path}")

    if not system_name:
        system_name = default_name

    # Column exclusion: explicit + auto-detected technical columns
    all_excluded: set[str] = set(exclude_columns or set())
    if not keep_technical:
        auto_technical = detect_technical_columns(tables)
        if auto_technical:
            logger.info(
                "Auto-excluding technical columns: %s (use --keep-technical to override)",
                ", ".join(sorted(auto_technical)),
            )
            all_excluded |= auto_technical
    if all_excluded:
        exclude_columns_from_tables(tables, all_excluded)

    if output_dir is None:
        from .hub_utils import resolve_hub_output_dir

        output_dir, hub_root = resolve_hub_output_dir(Path("integration") / "sources" / system_name)
        if hub_root is None:
            logger.warning(
                "Could not detect ontology-hub root (no ontology-hub/ or "
                "model/ontologies/ found). "
                "Writing to relative path: %s. "
                "Use --output to specify an explicit output directory.",
                output_dir,
            )

    # Issue #562: capture raw (pre-redaction) sample values for the alignment
    # LLM prompt, before write_source_dir's deep copy is redacted. A separate,
    # gitignored channel -- never changes what gets committed to the source dir.
    from .hub_utils import find_hub_root
    from .raw_samples import extract_raw_samples_from_tables, write_raw_samples

    write_raw_samples(find_hub_root(), system_name, extract_raw_samples_from_tables(tables))

    result_dir = write_source_dir(tables, system_name, output_dir, redact_pii=redact_pii)
    if return_count:
        sample_count = sum(1 for table in tables if table.get("sample_rows"))
        return result_dir, len(tables), sample_count, failures
    return result_dir
